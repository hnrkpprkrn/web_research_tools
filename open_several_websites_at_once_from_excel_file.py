# Import necessary libraries
import pandas as pd
from openpyxl import load_workbook
import webbrowser
import tkinter as tk
from tkinter import filedialog

def open_links_from_excel(root):
    """
    Function to open hyperlinks from an Excel file.

    Parameters:
    - root: Tkinter root window.

    This function hides the root window, prompts the user to select an Excel file,
    reads the data from the specified sheet and column, and opens each hyperlink in a new tab.
    """
    root.withdraw()  # Hide the root window
    file_path = filedialog.askopenfilename()  # Open file dialog to select file

    if file_path:
        sheet_name = 'Tabelle1'  # Name of the sheet containing the hyperlinks
        header_name = 'Link'  # Name of the column containing the hyperlinks
        workbook = load_workbook(filename=file_path)

        # Check if the specified sheet exists in the Excel file
        if sheet_name not in workbook.sheetnames:
            print(f"No worksheet named '{sheet_name}' found. Please make sure you have the correct sheet name.")
            return

        # Read the data from the specified sheet
        data = pd.read_excel(file_path, sheet_name=sheet_name)

        # Check if the specified column header exists in the data
        if header_name not in data.columns:
            print(f"No column named '{header_name}' found. Please make sure you have the correct column header.")
            return

        # Extract the hyperlinks from the specified column and open them in new tabs
        links = data[header_name]
        for link in links:
            if pd.notna(link):  # Check if the cell contains a valid hyperlink
                webbrowser.open_new_tab(link)

    root.quit()  # Close the Tkinter root window after processing

# Create the Tkinter root window
root = tk.Tk()
root.configure(background='black')  # Set background color to black

# Create a button to open hyperlinks
open_button = tk.Button(root, text="Open Links", bg="black", fg="white", command=lambda: open_links_from_excel(root))
open_button.pack()

root.mainloop()  # Start the Tkinter event loop
